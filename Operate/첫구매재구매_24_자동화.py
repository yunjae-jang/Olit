import pandas as pd
from openpyxl import load_workbook
################################얼라인랩#################################
workbook1 = load_workbook('C:/Users/김민경/파이썬파일/첫구매재구매/얼라인랩첫구매/첫구매고객수통합.xlsx')
workbook2 = load_workbook('C:/Users/김민경/파이썬파일/첫구매재구매/얼라인랩첫구매/회원첫구매고객수통합.xlsx')
workbook3 = load_workbook('C:/Users/김민경/파이썬파일/첫구매재구매/얼라인랩첫구매/첫구매통합.xlsx') 
workbook4 = load_workbook('C:/Users/김민경/파이썬파일/첫구매재구매/얼라인랩재구매/재구매통합.xlsx') 
sheet = workbook1.active
header_row = 1
num_columns = sheet.max_column
column_names = ['날짜', '첫구매고객수']
for col_idx, column_name in enumerate(column_names):
    cell = sheet.cell(row=header_row, column=col_idx+1)
    cell.value = column_name
workbook1.save('C:/Users/김민경/파이썬파일/첫구매재구매/얼라인랩첫구매/첫구매고객수통합.xlsx')
##
sheet = workbook2.active
header_row = 1
num_columns = sheet.max_column
column_names = ['날짜', '회원첫구매고객수']
for col_idx, column_name in enumerate(column_names):
    cell = sheet.cell(row=header_row, column=col_idx+1)
    cell.value = column_name
workbook2.save('C:/Users/김민경/파이썬파일/첫구매재구매/얼라인랩첫구매/회원첫구매고객수통합.xlsx')
##
sheet = workbook3.active
header_row = 1
num_columns = sheet.max_column
column_names = ['날짜','첫구매수','첫구매총액']
for col_idx, column_name in enumerate(column_names):
    cell = sheet.cell(row=header_row, column=col_idx+1)
    cell.value = column_name
workbook3.save('C:/Users/김민경/파이썬파일/첫구매재구매/얼라인랩첫구매/첫구매통합.xlsx')
##
sheet = workbook4.active
header_row = 1
num_columns = sheet.max_column
column_names = ['날짜', '재구매수','재구매총액']
for col_idx, column_name in enumerate(column_names):
    cell = sheet.cell(row=header_row, column=col_idx+1)
    cell.value = column_name
workbook4.save('C:/Users/김민경/파이썬파일/첫구매재구매/얼라인랩재구매/재구매통합.xlsx')

file1_path ='C:/Users/김민경/파이썬파일/첫구매재구매/얼라인랩첫구매/첫구매고객수통합.xlsx'
file2_path = 'C:/Users/김민경/파이썬파일/첫구매재구매/얼라인랩첫구매/회원첫구매고객수통합.xlsx'
file3_path = 'C:/Users/김민경/파이썬파일/첫구매재구매/얼라인랩첫구매/첫구매통합.xlsx' ##열 2개 있는 애 
file4_path = 'C:/Users/김민경/파이썬파일/첫구매재구매/얼라인랩재구매/재구매통합.xlsx' ##열 2개 있는 애
df1 = pd.read_excel(file1_path)
df2 = pd.read_excel(file2_path)
df3 = pd.read_excel(file3_path)
df4 = pd.read_excel(file4_path)
column_names = ['날짜', '첫구매고객수', '회원첫구매고객수','첫구매수', '첫구매총액','재구매수','재구매총액']

df1 = df1.sort_values(by='날짜')
df2 = df2.sort_values(by='날짜')
df3 = df3.sort_values(by='날짜')
df4 = df4.sort_values(by='날짜')

df_combined1 = pd.concat([df1['날짜'], df1['첫구매고객수'], df2['회원첫구매고객수'], df3['첫구매수'], df3['첫구매총액'], df4['재구매수'], df4['재구매총액']], axis=1, keys=column_names)

df_combined1.to_excel('C:/Users/김민경/파이썬파일/첫구매재구매/얼라인랩데이터통합.xlsx', index=False)

################################닥터아망#################################
workbook1 = load_workbook('C:/Users/김민경/파이썬파일/첫구매재구매/닥터아망첫구매/첫구매고객수통합.xlsx')
workbook2 = load_workbook('C:/Users/김민경/파이썬파일/첫구매재구매/닥터아망첫구매/회원첫구매고객수통합.xlsx')
workbook3 = load_workbook('C:/Users/김민경/파이썬파일/첫구매재구매/닥터아망첫구매/첫구매통합.xlsx') 
workbook4 = load_workbook('C:/Users/김민경/파이썬파일/첫구매재구매/닥터아망재구매/재구매통합.xlsx') 
sheet = workbook1.active
header_row = 1
num_columns = sheet.max_column
column_names = ['날짜', '첫구매고객수']
for col_idx, column_name in enumerate(column_names):
    cell = sheet.cell(row=header_row, column=col_idx+1)
    cell.value = column_name
workbook1.save('C:/Users/김민경/파이썬파일/첫구매재구매/닥터아망첫구매/첫구매고객수통합.xlsx')
##
sheet = workbook2.active
header_row = 1
num_columns = sheet.max_column
column_names = ['날짜', '회원첫구매고객수']
for col_idx, column_name in enumerate(column_names):
    cell = sheet.cell(row=header_row, column=col_idx+1)
    cell.value = column_name
workbook2.save('C:/Users/김민경/파이썬파일/첫구매재구매/닥터아망첫구매/회원첫구매고객수통합.xlsx')
##
sheet = workbook3.active
header_row = 1
num_columns = sheet.max_column
column_names = ['날짜','첫구매수','첫구매총액']
for col_idx, column_name in enumerate(column_names):
    cell = sheet.cell(row=header_row, column=col_idx+1)
    cell.value = column_name
workbook3.save('C:/Users/김민경/파이썬파일/첫구매재구매/닥터아망첫구매/첫구매통합.xlsx')
##
sheet = workbook4.active
header_row = 1
num_columns = sheet.max_column
column_names = ['날짜', '재구매수','재구매총액']
for col_idx, column_name in enumerate(column_names):
    cell = sheet.cell(row=header_row, column=col_idx+1)
    cell.value = column_name
workbook4.save('C:/Users/김민경/파이썬파일/첫구매재구매/닥터아망재구매/재구매통합.xlsx')

file1_path ='C:/Users/김민경/파이썬파일/첫구매재구매/닥터아망첫구매/첫구매고객수통합.xlsx'
file2_path = 'C:/Users/김민경/파이썬파일/첫구매재구매/닥터아망첫구매/회원첫구매고객수통합.xlsx'
file3_path = 'C:/Users/김민경/파이썬파일/첫구매재구매/닥터아망첫구매/첫구매통합.xlsx' ##열 2개 있는 애 
file4_path = 'C:/Users/김민경/파이썬파일/첫구매재구매/닥터아망재구매/재구매통합.xlsx' ##열 2개 있는 애

df1 = pd.read_excel(file1_path)
df2 = pd.read_excel(file2_path)
df3 = pd.read_excel(file3_path)
df4 = pd.read_excel(file4_path)

column_names = ['날짜', '첫구매고객수', '회원첫구매고객수','첫구매수', '첫구매총액','재구매수','재구매총액']

df1 = df1.sort_values(by='날짜')
df2 = df2.sort_values(by='날짜')
df3 = df3.sort_values(by='날짜')
df4 = df4.sort_values(by='날짜')

df_combined1 = pd.concat([df1['날짜'], df1['첫구매고객수'], df2['회원첫구매고객수'], df3['첫구매수'], df3['첫구매총액'], df4['재구매수'], df4['재구매총액']], axis=1, keys=column_names)

df_combined1.to_excel('C:/Users/김민경/파이썬파일/첫구매재구매/닥터아망데이터통합.xlsx', index=False)
################################와이브닝#################################
workbook1 = load_workbook('C:/Users/김민경/파이썬파일/첫구매재구매/와이브닝첫구매/첫구매고객수통합.xlsx')
workbook2 = load_workbook('C:/Users/김민경/파이썬파일/첫구매재구매/와이브닝첫구매/회원첫구매고객수통합.xlsx')
workbook3 = load_workbook('C:/Users/김민경/파이썬파일/첫구매재구매/와이브닝첫구매/첫구매통합.xlsx') 
workbook4 = load_workbook('C:/Users/김민경/파이썬파일/첫구매재구매/와이브닝재구매/재구매통합.xlsx') 
sheet = workbook1.active
header_row = 1
num_columns = sheet.max_column
column_names = ['날짜', '첫구매고객수']
for col_idx, column_name in enumerate(column_names):
    cell = sheet.cell(row=header_row, column=col_idx+1)
    cell.value = column_name
workbook1.save('C:/Users/김민경/파이썬파일/첫구매재구매/와이브닝첫구매/첫구매고객수통합.xlsx')
##
sheet = workbook2.active
header_row = 1
num_columns = sheet.max_column
column_names = ['날짜', '회원첫구매고객수']
for col_idx, column_name in enumerate(column_names):
    cell = sheet.cell(row=header_row, column=col_idx+1)
    cell.value = column_name
workbook2.save('C:/Users/김민경/파이썬파일/첫구매재구매/와이브닝첫구매/회원첫구매고객수통합.xlsx')
##
sheet = workbook3.active
header_row = 1
num_columns = sheet.max_column
column_names = ['날짜','첫구매수','첫구매총액']
for col_idx, column_name in enumerate(column_names):
    cell = sheet.cell(row=header_row, column=col_idx+1)
    cell.value = column_name
workbook3.save('C:/Users/김민경/파이썬파일/첫구매재구매/와이브닝첫구매/첫구매통합.xlsx')
##
sheet = workbook4.active
header_row = 1
num_columns = sheet.max_column
column_names = ['날짜', '재구매수','재구매총액']
for col_idx, column_name in enumerate(column_names):
    cell = sheet.cell(row=header_row, column=col_idx+1)
    cell.value = column_name
workbook4.save('C:/Users/김민경/파이썬파일/첫구매재구매/와이브닝재구매/재구매통합.xlsx')

file1_path ='C:/Users/김민경/파이썬파일/첫구매재구매/와이브닝첫구매/첫구매고객수통합.xlsx'
file2_path = 'C:/Users/김민경/파이썬파일/첫구매재구매/와이브닝첫구매/회원첫구매고객수통합.xlsx'
file3_path = 'C:/Users/김민경/파이썬파일/첫구매재구매/와이브닝첫구매/첫구매통합.xlsx' ##열 2개 있는 애 
file4_path = 'C:/Users/김민경/파이썬파일/첫구매재구매/와이브닝재구매/재구매통합.xlsx' ##열 2개 있는 애

df1 = pd.read_excel(file1_path)
df2 = pd.read_excel(file2_path)
df3 = pd.read_excel(file3_path)
df4 = pd.read_excel(file4_path)

column_names = ['날짜', '첫구매고객수', '회원첫구매고객수','첫구매수', '첫구매총액','재구매수','재구매총액']

df1 = df1.sort_values(by='날짜')
df2 = df2.sort_values(by='날짜')
df3 = df3.sort_values(by='날짜')
df4 = df4.sort_values(by='날짜')

df_combined1 = pd.concat([df1['날짜'], df1['첫구매고객수'], df2['회원첫구매고객수'], df3['첫구매수'], df3['첫구매총액'], df4['재구매수'], df4['재구매총액']], axis=1, keys=column_names)

df_combined1.to_excel('C:/Users/김민경/파이썬파일/첫구매재구매/와이브닝데이터통합.xlsx', index=False)
################################셀올로지#################################
workbook1 = load_workbook('C:/Users/김민경/파이썬파일/첫구매재구매/셀올로지첫구매/첫구매고객수통합.xlsx')
workbook2 = load_workbook('C:/Users/김민경/파이썬파일/첫구매재구매/셀올로지첫구매/회원첫구매고객수통합.xlsx')
workbook3 = load_workbook('C:/Users/김민경/파이썬파일/첫구매재구매/셀올로지첫구매/첫구매통합.xlsx') 
workbook4 = load_workbook('C:/Users/김민경/파이썬파일/첫구매재구매/셀올로지재구매/재구매통합.xlsx') 
sheet = workbook1.active
header_row = 1
num_columns = sheet.max_column
column_names = ['날짜', '첫구매고객수']
for col_idx, column_name in enumerate(column_names):
    cell = sheet.cell(row=header_row, column=col_idx+1)
    cell.value = column_name
workbook1.save('C:/Users/김민경/파이썬파일/첫구매재구매/셀올로지첫구매/첫구매고객수통합.xlsx')
##
sheet = workbook2.active
header_row = 1
num_columns = sheet.max_column
column_names = ['날짜', '회원첫구매고객수']
for col_idx, column_name in enumerate(column_names):
    cell = sheet.cell(row=header_row, column=col_idx+1)
    cell.value = column_name
workbook2.save('C:/Users/김민경/파이썬파일/첫구매재구매/셀올로지첫구매/회원첫구매고객수통합.xlsx')
##
sheet = workbook3.active
header_row = 1
num_columns = sheet.max_column
column_names = ['날짜','첫구매수','첫구매총액']
for col_idx, column_name in enumerate(column_names):
    cell = sheet.cell(row=header_row, column=col_idx+1)
    cell.value = column_name
workbook3.save('C:/Users/김민경/파이썬파일/첫구매재구매/셀올로지첫구매/첫구매통합.xlsx')
##
sheet = workbook4.active
header_row = 1
num_columns = sheet.max_column
column_names = ['날짜', '재구매수','재구매총액']
for col_idx, column_name in enumerate(column_names):
    cell = sheet.cell(row=header_row, column=col_idx+1)
    cell.value = column_name
workbook4.save('C:/Users/김민경/파이썬파일/첫구매재구매/셀올로지재구매/재구매통합.xlsx')

file1_path ='C:/Users/김민경/파이썬파일/첫구매재구매/셀올로지첫구매/첫구매고객수통합.xlsx'
file2_path = 'C:/Users/김민경/파이썬파일/첫구매재구매/셀올로지첫구매/회원첫구매고객수통합.xlsx'
file3_path = 'C:/Users/김민경/파이썬파일/첫구매재구매/셀올로지첫구매/첫구매통합.xlsx' ##열 2개 있는 애 
file4_path = 'C:/Users/김민경/파이썬파일/첫구매재구매/셀올로지재구매/재구매통합.xlsx' ##열 2개 있는 애

df1 = pd.read_excel(file1_path)
df2 = pd.read_excel(file2_path)
df3 = pd.read_excel(file3_path)
df4 = pd.read_excel(file4_path)

column_names = ['날짜', '첫구매고객수', '회원첫구매고객수','첫구매수', '첫구매총액','재구매수','재구매총액']

df1 = df1.sort_values(by='날짜')
df2 = df2.sort_values(by='날짜')
df3 = df3.sort_values(by='날짜')
df4 = df4.sort_values(by='날짜')

df_combined1 = pd.concat([df1['날짜'], df1['첫구매고객수'], df2['회원첫구매고객수'], df3['첫구매수'], df3['첫구매총액'], df4['재구매수'], df4['재구매총액']], axis=1, keys=column_names)

df_combined1.to_excel('C:/Users/김민경/파이썬파일/첫구매재구매/셀올로지데이터통합.xlsx', index=False)
